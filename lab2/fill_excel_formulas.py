from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(__file__).resolve().parent / "lab2_Student.xlsx"
SHEET_NAME = "Sheet1"
START_ROW = 11
END_ROW = 110


def apply_number_formats(ws):
    for row in range(START_ROW, END_ROW + 1):
        ws[f"B{row}"].number_format = "0.000000"
        ws[f"C{row}"].number_format = "0.000000"
        ws[f"D{row}"].number_format = "0.000000"
        ws[f"E{row}"].number_format = "0.000000"
        ws[f"F{row}"].number_format = "0.000000"
        ws[f"G{row}"].number_format = "0.000000"
        ws[f"H{row}"].number_format = "0.000000"

    for cell in ["B8", "D9", "F3", "F4", "F5", "F7", "F8", "F9", "H3", "H4", "H5", "H7"]:
        ws[cell].number_format = "0.000000"


def fill_formulas(ws):
    ws["B8"] = "=B3/B5"

    for row in range(START_ROW, END_ROW + 1):
        x = f"A{row}"
        q = f"B{row}"
        lam_x = f"C{row}"
        mu_x = f"D{row}"
        t_x = f"E{row}"
        p_x = f"F{row}"
        lam_px = f"G{row}"
        q_px = f"H{row}"

        ws[q] = f'=IF({x}<$B$4,0,{x}-$B$4)'
        ws[lam_x] = (
            f'=IF({x}>$B$6,0,'
            f'IF($B$7=0,$B$8,IF({x}<=$B$7,$B$8*($B$7-{x}),0)))'
        )
        ws[mu_x] = f'=IF({x}>$B$6,0,IF({x}>$B$4,$B$4/$B$5,{x}/$B$5))'

        if row == START_ROW:
            ws[t_x] = "=1"
        else:
            prev_lam_x = f"C{row - 1}"
            prev_t_x = f"E{row - 1}"
            ws[t_x] = f'=IF({x}>$B$6,0,IF({mu_x}<>0,{prev_t_x}*({prev_lam_x}/{mu_x}),0))'

        ws[p_x] = f'=IF({x}>$B$6,0,{t_x}/$D$9)'
        ws[lam_px] = f"={lam_x}*{p_x}"
        ws[q_px] = f"={q}*{p_x}"

    ws["D9"] = f"=SUM(E{START_ROW}:E{END_ROW})"
    ws["F9"] = f"=SUM(G{START_ROW}:G{END_ROW})"

    ws["F3"] = f'=IF($F$9=0,0,SUMIFS($G${START_ROW}:$G${END_ROW},$A${START_ROW}:$A${END_ROW},"<"&$B$4)/$F$9)'
    ws["F4"] = (
        f'=IF($F$9=0,0,'
        f'SUMIFS($G${START_ROW}:$G${END_ROW},$A${START_ROW}:$A${END_ROW},">="&$B$4,$A${START_ROW}:$A${END_ROW},"<"&$B$6)/$F$9)'
    )
    ws["F5"] = f'=IF($F$9=0,0,SUMIFS($G${START_ROW}:$G${END_ROW},$A${START_ROW}:$A${END_ROW},$B$6)/$F$9)'
    ws["F7"] = f"=SUM(H{START_ROW}:H{END_ROW})"
    ws["F8"] = f"=SUMPRODUCT(A{START_ROW}:A{END_ROW},F{START_ROW}:F{END_ROW})"

    ws["H3"] = f'=SUMIFS($F${START_ROW}:$F${END_ROW},$A${START_ROW}:$A${END_ROW},$B$6)'
    ws["H4"] = (
        f'=SUMIFS($F${START_ROW}:$F${END_ROW},$A${START_ROW}:$A${END_ROW},">="&$B$4,$A${START_ROW}:$A${END_ROW},"<"&$B$6)'
    )
    ws["H5"] = f'=SUMIFS($F${START_ROW}:$F${END_ROW},$A${START_ROW}:$A${END_ROW},"<"&$B$4)'
    ws["H7"] = (
        f'=IF(SUMIFS($F${START_ROW}:$F${END_ROW},$A${START_ROW}:$A${END_ROW},">"&$B$4)=0,0,'
        f'$F$7/SUMIFS($F${START_ROW}:$F${END_ROW},$A${START_ROW}:$A${END_ROW},">"&$B$4))'
    )


def main():
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    fill_formulas(ws)
    apply_number_formats(ws)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    wb.save(WORKBOOK_PATH)
    print(f"Updated formulas in {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
